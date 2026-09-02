from __future__ import annotations

import os
import tempfile
from collections.abc import Iterable, Mapping
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lead_enrichment.models import (
    ChannelType,
    CompanyInput,
    KonturImportResult,
    SourceReference,
    SourceResult,
)

EXCEL_CELL_LIMIT = 32_767
REQUIRED_EXPORT_HEADERS: dict[str, tuple[str, ...]] = {
    "Компании": ("Строка", "Наименование", "ИНН", "Тип", "Статус обработки"),
    "Контакты": ("ИНН", "Компания", "ФИО", "Должность", "Роль"),
    "Источники": ("ИНН", "Источник", "Тип источника", "Дата сбора"),
    "Журнал": ("ИНН", "Этап", "Результат", "Код причины"),
    "Сводка": ("Показатель", "Значение"),
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def export_enrichment_workbook(
    import_result: KonturImportResult,
    output_file_path: Path,
    *,
    site_results: Mapping[str, SourceResult] | None = None,
) -> Path:
    target = Path(output_file_path)
    if target.suffix.casefold() != ".xlsx":
        raise ValueError("Итоговый файл должен иметь расширение .xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    site_results = site_results or {}

    workbook = Workbook()
    try:
        workbook.remove(workbook.active)
        companies_sheet = workbook.create_sheet("Компании")
        contacts_sheet = workbook.create_sheet("Контакты")
        sources_sheet = workbook.create_sheet("Источники")
        journal_sheet = workbook.create_sheet("Журнал")
        summary_sheet = workbook.create_sheet("Сводка", 0)

        _write_companies(companies_sheet, import_result.companies, site_results)
        _write_contacts(contacts_sheet, import_result.companies, site_results)
        _write_sources(sources_sheet, import_result.companies, site_results)
        _write_journal(journal_sheet, import_result.companies, site_results)
        _write_summary(summary_sheet, import_result)

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        _atomic_save_workbook(workbook, target)
    finally:
        workbook.close()
    return target


def sanitize_excel_text(value: str, *, protect_formulas: bool = True) -> str:
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    first_visible = cleaned.lstrip(" \t\r\n")[:1]
    if protect_formulas and first_visible in {"=", "+", "-", "@"}:
        cleaned = f"'{cleaned}"
    return cleaned[:EXCEL_CELL_LIMIT]


def _write_companies(sheet, companies: list[CompanyInput], site_results: Mapping[str, SourceResult]) -> None:
    headers = [
        "Строка",
        "Наименование",
        "ИНН",
        "Тип",
        "Статус обработки",
        "Регион",
        "Сегмент",
        "Сайт",
        "Email компании",
        "Телефоны компании",
        "Людей найдено",
        "Результат сайта",
        "Причина сайта",
    ]
    sheet.append(headers)
    for company in companies:
        site_result = site_results.get(company.inn)
        channels = _merged_channels(company, site_result)
        sheet.append([
            _safe_value(company.input_row_id),
            _safe_value(company.legal_name),
            company.inn,
            company.entity_type.value,
            "IMPORTED",
            _safe_value(company.region),
            _safe_value(company.segment_name),
            _safe_value(company.website),
            _safe_value(_join_channel_values(channels, ChannelType.EMAIL)),
            _safe_value(_join_channel_values(channels, ChannelType.PHONE)),
            len(company.initial_people) + (len(site_result.person_contacts) if site_result else 0),
            site_result.outcome.value if site_result else None,
            _safe_value(site_result.reason_message if site_result else None),
        ])
        if company.website and _is_safe_http_hyperlink(company.website):
            cell = sheet.cell(row=sheet.max_row, column=8)
            cell.hyperlink = company.website
            cell.style = "Hyperlink"
    _format_tabular_sheet(sheet, widths=[20, 45, 15, 24, 20, 24, 30, 38, 45, 35, 15, 18, 50])


def _write_contacts(sheet, companies: list[CompanyInput], site_results: Mapping[str, SourceResult]) -> None:
    headers = [
        "ИНН",
        "Компания",
        "ФИО",
        "Должность",
        "Роль",
        "Email компании",
        "Телефоны компании",
        "Соцсети компании",
        "Уверенность",
        "Источник человека",
        "Дата сбора",
    ]
    sheet.append(headers)
    for company in companies:
        site_result = site_results.get(company.inn)
        channels = _merged_channels(company, site_result)
        people = list(company.initial_people)
        if site_result:
            people.extend(site_result.person_contacts)
        if not people:
            sheet.append([
                company.inn,
                _safe_value(company.legal_name),
                None,
                None,
                None,
                _safe_value(_join_channel_values(channels, ChannelType.EMAIL)),
                _safe_value(_join_channel_values(channels, ChannelType.PHONE)),
                _safe_value(_join_channel_values(channels, ChannelType.SOCIAL)),
                None,
                None,
                None,
            ])
            continue
        for person in people:
            source_names = _unique(ref.source_name for ref in person.source_refs)
            collected_dates = [ref.collected_at for ref in person.source_refs]
            sheet.append([
                company.inn,
                _safe_value(company.legal_name),
                _safe_value(person.full_name),
                _safe_value(person.job_title),
                person.normalized_role.value,
                _safe_value(_join_channel_values(channels, ChannelType.EMAIL)),
                _safe_value(_join_channel_values(channels, ChannelType.PHONE)),
                _safe_value(_join_channel_values(channels, ChannelType.SOCIAL)),
                person.confidence_score,
                _safe_value("; ".join(source_names)),
                _excel_datetime(min(collected_dates)) if collected_dates else None,
            ])
    _format_tabular_sheet(sheet, widths=[15, 45, 34, 32, 18, 45, 35, 45, 14, 28, 22])
    for cell in sheet["K"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm:ss"


def _write_sources(sheet, companies: list[CompanyInput], site_results: Mapping[str, SourceResult]) -> None:
    headers = [
        "ИНН",
        "Источник",
        "Тип источника",
        "Дата сбора",
        "Надёжность",
        "Локатор",
        "URL",
    ]
    sheet.append(headers)
    seen: set[tuple[str, str, str, str | None]] = set()
    for company in companies:
        refs = list(_company_source_refs(company))
        site_result = site_results.get(company.inn)
        if site_result:
            for channel in site_result.company_channels:
                refs.extend(channel.source_refs)
            for person in site_result.person_contacts:
                refs.extend(person.source_refs)
        for ref in refs:
            key = (company.inn, ref.source_id, ref.locator, ref.url)
            if key in seen:
                continue
            seen.add(key)
            sheet.append([
                company.inn,
                _safe_value(ref.source_name),
                ref.source_kind.value,
                _excel_datetime(ref.collected_at),
                ref.reliability,
                _safe_value(ref.locator),
                _safe_value(ref.url),
            ])
            if ref.url and _is_safe_http_hyperlink(ref.url):
                cell = sheet.cell(row=sheet.max_row, column=7)
                cell.hyperlink = ref.url
                cell.style = "Hyperlink"
    _format_tabular_sheet(sheet, widths=[15, 28, 24, 22, 14, 34, 55])
    for cell in sheet["D"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm:ss"


def _write_journal(sheet, companies: list[CompanyInput], site_results: Mapping[str, SourceResult]) -> None:
    headers = [
        "ИНН",
        "Этап",
        "Результат",
        "Код причины",
        "Причина",
        "Причина продолжения",
        "Запросов",
        "Страниц проверено",
        "Длительность, мс",
    ]
    sheet.append(headers)
    for company in companies:
        sheet.append([
            company.inn,
            "kontur_import",
            "FOUND",
            "KONTUR_ROW_IMPORTED",
            "Строка Контур успешно импортирована",
            None,
            0,
            0,
            0,
        ])
        site_result = site_results.get(company.inn)
        if site_result:
            sheet.append([
                company.inn,
                site_result.source_id,
                site_result.outcome.value,
                site_result.reason_code,
                _safe_value(site_result.reason_message),
                _safe_value(site_result.continue_reason),
                site_result.metrics.request_count,
                site_result.metrics.checked_page_count,
                site_result.metrics.duration_ms,
            ])
    _format_tabular_sheet(sheet, widths=[15, 24, 16, 30, 55, 55, 12, 18, 18])


def _write_summary(sheet, import_result: KonturImportResult) -> None:
    sheet["A1"] = "AdBeam Person Parser — сводка"
    sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    sheet.merge_cells("A1:B1")
    sheet.append([])
    sheet.append(["Показатель", "Значение"])
    metrics = [
        ("Исходный файл", import_result.summary.source_file_name),
        ("Лист", import_result.summary.sheet_name),
        ("Строк в импорте", import_result.summary.total_rows),
        ("Импортировано", import_result.summary.imported_rows),
        ("Пропущено", import_result.summary.skipped_rows),
        ("Дубликатов ИНН", import_result.summary.duplicate_inn_rows),
        ("Юрлиц", import_result.summary.legal_entities),
        ("ИП", import_result.summary.individual_entrepreneurs),
        ("Есть исходный человек", import_result.summary.rows_with_initial_person),
        ("Есть email", import_result.summary.rows_with_emails),
        ("Есть телефон", import_result.summary.rows_with_phones),
        ("Есть сайт", import_result.summary.rows_with_website),
    ]
    for label, value in metrics:
        sheet.append([_safe_value(label), _safe_value(value)])
    sheet["A3"].fill = HEADER_FILL
    sheet["B3"].fill = HEADER_FILL
    sheet["A3"].font = HEADER_FONT
    sheet["B3"].font = HEADER_FONT
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 50
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    for row in range(4, sheet.max_row + 1):
        if row % 2 == 0:
            sheet.cell(row=row, column=1).fill = SUBTITLE_FILL


def _format_tabular_sheet(sheet, widths: list[int]) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 60)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def _atomic_save_workbook(workbook: Workbook, target: Path) -> None:
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{target.stem}.",
        suffix=".tmp.xlsx",
        dir=target.parent,
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        try:
            workbook.save(temp_path)
        finally:
            workbook.close()
        _validate_export_workbook(temp_path)
        os.replace(temp_path, target)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _validate_export_workbook(path: Path) -> None:
    validation = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        for sheet_name, required_headers in REQUIRED_EXPORT_HEADERS.items():
            if sheet_name not in validation.sheetnames:
                raise ValueError(f"В экспортированном файле отсутствует лист {sheet_name}")
            sheet = validation[sheet_name]
            header_row = 3 if sheet_name == "Сводка" else 1
            observed = {
                str(cell.value).strip()
                for cell in sheet[header_row]
                if cell.value is not None
            }
            missing = set(required_headers) - observed
            if missing:
                raise ValueError(f"В листе {sheet_name} отсутствуют обязательные заголовки")
    finally:
        validation.close()


def _merged_channels(company: CompanyInput, site_result: SourceResult | None):
    channels = list(company.company_channels)
    if site_result:
        channels.extend(site_result.company_channels)
    result = []
    seen: set[tuple[str, str]] = set()
    for channel in channels:
        key = (channel.channel_type.value, channel.value.casefold())
        if key in seen:
            continue
        seen.add(key)
        result.append(channel)
    return result


def _join_channel_values(channels, channel_type: ChannelType) -> str | None:
    values = _unique(
        channel.value for channel in channels if channel.channel_type == channel_type
    )
    return "; ".join(values) or None


def _company_source_refs(company: CompanyInput) -> Iterable[SourceReference]:
    yield from company.source_refs
    for channel in company.company_channels:
        yield from channel.source_refs
    for person in company.initial_people:
        yield from person.source_refs


def _unique(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _safe_value(value):
    if isinstance(value, datetime):
        return _excel_datetime(value)
    if value is None or isinstance(value, (int, float, bool, date)):
        return value
    return sanitize_excel_text(str(value))


def _excel_datetime(value: datetime) -> datetime:
    if value.tzinfo is None or value.utcoffset() is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _is_safe_http_hyperlink(value: str) -> bool:
    try:
        parsed = urlsplit(value)
    except ValueError:
        return False
    return bool(
        parsed.scheme.casefold() in {"http", "https"}
        and parsed.hostname
        and not parsed.username
        and not parsed.password
    )
