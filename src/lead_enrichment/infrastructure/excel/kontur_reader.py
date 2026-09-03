from __future__ import annotations

import hashlib
import re
import warnings
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from uuid import UUID, uuid5

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from lead_enrichment.engine.normalization import (
    extract_ip_owner_name,
    infer_entity_type,
    normalize_http_url,
    normalize_identifier,
    normalize_inn,
    normalize_person_name,
    split_emails,
    split_lines,
    split_phones,
)
from lead_enrichment.models import (
    ChannelScope,
    ChannelType,
    CompanyFinancials,
    CompanyInput,
    ContactChannel,
    ContactRole,
    EntityType,
    ImportIssue,
    ImportIssueSeverity,
    KonturImportResult,
    KonturImportSummary,
    PersonContact,
    SourceKind,
    SourceReference,
)

CONTACT_NAMESPACE = UUID("1dbb5ff7-138c-4aa4-b7f9-9421bcf444be")
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
DEFAULT_SHEET_NAME = "Контрагенты"

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "legal_name": ("Наименование",),
    "inn": ("ИНН",),
    "kpp": ("КПП",),
    "ogrn": ("ОГРН", "ОГРНИП"),
    "registration_date": ("Дата регистрации",),
    "address": ("Адрес",),
    "region": ("Регион регистрации", "Регион"),
    "operating_status": ("Статус",),
    "msp_category": ("Реестр МСП",),
    "focus_url": ("Карточка в Фокусе",),
    "manager_name": ("ФИО руководителя",),
    "manager_title": ("Должность руководителя",),
    "phones": ("Номер телефона", "Телефон", "Телефоны"),
    "emails": ("Электронная почта", "Email", "E-mail"),
    "website": ("Ссылка на сайт", "Сайт", "Веб-сайт"),
    "primary_activity": ("Основной вид деятельности",),
    "other_activities": ("Другие виды деятельности",),
    "licenses": ("Полученные лицензии",),
    "revenue": ("Выручка",),
    "balance": ("Баланс",),
    "net_profit_loss": ("Чистая прибыль/ убыток", "Чистая прибыль/убыток"),
    "arbitration_defendant": ("Арбитраж (ответчик)",),
    "employee_count": ("Количество сотрудников",),
    "branches": ("Филиалы",),
    "branch_count": ("Количество филиалов",),
    "source_label": ("Источник",),
    "segment_name": ("Название сегмента",),
}
REQUIRED_COLUMNS = {"legal_name", "inn"}


class KonturWorkbookError(ValueError):
    """The workbook cannot be interpreted as a supported Kontur export."""


class MissingRequiredColumnsError(KonturWorkbookError):
    def __init__(self, missing_columns: list[str]) -> None:
        self.missing_columns = missing_columns
        super().__init__(f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}")


def read_kontur_workbook(
    file_path: Path,
    *,
    sheet_name: str | None = None,
    collected_at: datetime | None = None,
) -> KonturImportResult:
    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError("Файл выгрузки Контур не найден")
    if path.suffix.casefold() not in SUPPORTED_SUFFIXES:
        raise KonturWorkbookError("Поддерживаются только файлы .xlsx и .xlsm")

    collected_at = collected_at or datetime.now(timezone.utc)
    if collected_at.tzinfo is None or collected_at.utcoffset() is None:
        raise ValueError("collected_at must be timezone-aware")

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style.*")
        workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)

    try:
        worksheet = _select_worksheet(workbook, sheet_name)
        column_map = _build_column_map(worksheet)
        missing = sorted(REQUIRED_COLUMNS - column_map.keys())
        if missing:
            readable = [COLUMN_ALIASES[name][0] for name in missing]
            raise MissingRequiredColumnsError(readable)

        companies: list[CompanyInput] = []
        issues: list[ImportIssue] = []
        seen_inn: set[str] = set()
        counters = _ImportCounters()

        for row_index in range(2, worksheet.max_row + 1):
            if _row_is_blank(worksheet, row_index):
                counters.blank_rows += 1
                continue

            legal_name = _text(_cell_value(worksheet, row_index, column_map, "legal_name"))
            inn = normalize_inn(_cell_value(worksheet, row_index, column_map, "inn"))
            if not legal_name:
                issues.append(_issue(row_index, "MISSING_LEGAL_NAME", "Не заполнено наименование компании"))
                continue
            if inn is None:
                issues.append(_issue(row_index, "INVALID_INN", "ИНН отсутствует или не прошёл проверку"))
                continue

            if inn in seen_inn:
                counters.duplicate_inn_rows += 1
                issues.append(
                    ImportIssue(
                        row_index=row_index,
                        code="DUPLICATE_INN",
                        message="ИНН уже встречался в этой выгрузке; повторная строка пропущена",
                        severity=ImportIssueSeverity.WARNING,
                    )
                )
                continue
            seen_inn.add(inn)

            entity_type = infer_entity_type(inn)
            counters.legal_entities += int(entity_type == EntityType.LEGAL_ENTITY)
            counters.individual_entrepreneurs += int(
                entity_type == EntityType.INDIVIDUAL_ENTREPRENEUR
            )

            focus_url = normalize_http_url(
                _cell_or_hyperlink(worksheet, row_index, column_map, "focus_url")
            )
            website = normalize_http_url(
                _cell_or_hyperlink(worksheet, row_index, column_map, "website")
            )
            source_label = _text(_cell_value(worksheet, row_index, column_map, "source_label"))
            source_ref = SourceReference(
                source_id="kontur_import",
                source_kind=SourceKind.KONTUR_EXPORT,
                source_name=source_label or "Контур.Продажи",
                locator=f"{worksheet.title}!{row_index}",
                collected_at=collected_at,
                url=focus_url,
                reliability=90,
            )

            emails = split_emails(_cell_value(worksheet, row_index, column_map, "emails"))
            phones = split_phones(_cell_value(worksheet, row_index, column_map, "phones"))
            counters.rows_with_emails += int(bool(emails))
            counters.rows_with_phones += int(bool(phones))
            counters.rows_with_website += int(website is not None)
            channels = [
                ContactChannel(
                    channel_type=ChannelType.EMAIL,
                    value=email,
                    scope=ChannelScope.COMPANY,
                    source_refs=[source_ref],
                )
                for email in emails
            ]
            channels.extend(
                ContactChannel(
                    channel_type=ChannelType.PHONE,
                    value=phone,
                    scope=ChannelScope.COMPANY,
                    source_refs=[source_ref],
                )
                for phone in phones
            )

            manager_name = normalize_person_name(
                _cell_value(worksheet, row_index, column_map, "manager_name")
            )
            manager_title = _text(_cell_value(worksheet, row_index, column_map, "manager_title"))
            role = ContactRole.LEADER
            confidence = 90
            if not manager_name and entity_type == EntityType.INDIVIDUAL_ENTREPRENEUR:
                manager_name = extract_ip_owner_name(legal_name)
                manager_title = manager_title or "Индивидуальный предприниматель"
                role = ContactRole.OWNER
                confidence = 88

            people: list[PersonContact] = []
            if manager_name:
                counters.rows_with_initial_person += 1
                people.append(
                    PersonContact(
                        contact_id=_contact_id(inn, manager_name, role),
                        full_name=manager_name,
                        job_title=manager_title,
                        normalized_role=role,
                        confidence_score=confidence,
                        source_refs=[source_ref],
                    )
                )

            companies.append(
                CompanyInput(
                    input_row_id=f"{worksheet.title}:{row_index}",
                    legal_name=legal_name,
                    inn=inn,
                    kpp=normalize_identifier(
                        _cell_value(worksheet, row_index, column_map, "kpp"), {9}
                    ),
                    ogrn=normalize_identifier(
                        _cell_value(worksheet, row_index, column_map, "ogrn"), {13, 15}
                    ),
                    entity_type=entity_type,
                    registration_date=_date_value(
                        _cell_value(worksheet, row_index, column_map, "registration_date")
                    ),
                    address=_text(_cell_value(worksheet, row_index, column_map, "address")),
                    region=_text(_cell_value(worksheet, row_index, column_map, "region")),
                    operating_status=_text(
                        _cell_value(worksheet, row_index, column_map, "operating_status")
                    ),
                    msp_category=_text(
                        _cell_value(worksheet, row_index, column_map, "msp_category")
                    ),
                    focus_url=focus_url,
                    website=website,
                    primary_activity=_text(
                        _cell_value(worksheet, row_index, column_map, "primary_activity")
                    ),
                    other_activities=split_lines(
                        _cell_value(worksheet, row_index, column_map, "other_activities")
                    ),
                    licenses=split_lines(
                        _cell_value(worksheet, row_index, column_map, "licenses")
                    ),
                    financials=CompanyFinancials(
                        revenue=_text(_cell_value(worksheet, row_index, column_map, "revenue")),
                        balance=_text(_cell_value(worksheet, row_index, column_map, "balance")),
                        net_profit_loss=_text(
                            _cell_value(worksheet, row_index, column_map, "net_profit_loss")
                        ),
                        arbitration_defendant=_text(
                            _cell_value(worksheet, row_index, column_map, "arbitration_defendant")
                        ),
                    ),
                    employee_count=_non_negative_int(
                        _cell_value(worksheet, row_index, column_map, "employee_count")
                    ),
                    branches=split_lines(
                        _cell_value(worksheet, row_index, column_map, "branches")
                    ),
                    branch_count=_non_negative_int(
                        _cell_value(worksheet, row_index, column_map, "branch_count")
                    ),
                    source_label=source_label,
                    segment_name=_text(
                        _cell_value(worksheet, row_index, column_map, "segment_name")
                    ),
                    company_channels=channels,
                    initial_people=people,
                    source_refs=[source_ref],
                )
            )

        total_rows = max(worksheet.max_row - 1, 0)
        summary = KonturImportSummary(
            source_file_name=path.name,
            source_sha256=_sha256(path),
            sheet_name=worksheet.title,
            total_rows=total_rows,
            imported_rows=len(companies),
            skipped_rows=total_rows - counters.blank_rows - len(companies),
            blank_rows=counters.blank_rows,
            duplicate_inn_rows=counters.duplicate_inn_rows,
            legal_entities=counters.legal_entities,
            individual_entrepreneurs=counters.individual_entrepreneurs,
            rows_with_initial_person=counters.rows_with_initial_person,
            rows_with_phones=counters.rows_with_phones,
            rows_with_emails=counters.rows_with_emails,
            rows_with_website=counters.rows_with_website,
            issues=issues,
        )
        return KonturImportResult(summary=summary, companies=companies)
    finally:
        workbook.close()


class _ImportCounters:
    def __init__(self) -> None:
        self.blank_rows = 0
        self.duplicate_inn_rows = 0
        self.legal_entities = 0
        self.individual_entrepreneurs = 0
        self.rows_with_initial_person = 0
        self.rows_with_phones = 0
        self.rows_with_emails = 0
        self.rows_with_website = 0


def _select_worksheet(workbook: Any, requested_name: str | None):
    if requested_name:
        if requested_name not in workbook.sheetnames:
            raise KonturWorkbookError("Запрошенный лист отсутствует в книге")
        return workbook[requested_name]
    if DEFAULT_SHEET_NAME in workbook.sheetnames:
        return workbook[DEFAULT_SHEET_NAME]
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    raise KonturWorkbookError("Не удалось однозначно выбрать лист Контур")


def _build_column_map(worksheet) -> dict[str, int]:
    observed: dict[str, int] = {}
    for column_index, cell in enumerate(worksheet[1], start=1):
        normalized = _normalize_header(cell.value)
        if not normalized:
            continue
        if normalized in observed:
            raise KonturWorkbookError("В строке заголовков есть дублирующиеся колонки")
        observed[normalized] = column_index

    result: dict[str, int] = {}
    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            column_index = observed.get(_normalize_header(alias))
            if column_index is not None:
                result[field_name] = column_index
                break
    return result


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    normalized = re.sub(r"\s+", " ", str(value)).strip().casefold()
    return re.sub(r"\s*/\s*", "/", normalized)


def _row_is_blank(worksheet, row_index: int) -> bool:
    return not any(
        cell.value is not None and str(cell.value).strip() for cell in worksheet[row_index]
    )


def _cell_value(worksheet, row_index: int, column_map: dict[str, int], field_name: str):
    column_index = column_map.get(field_name)
    return worksheet.cell(row=row_index, column=column_index).value if column_index else None


def _cell_or_hyperlink(
    worksheet,
    row_index: int,
    column_map: dict[str, int],
    field_name: str,
) -> object:
    column_index = column_map.get(field_name)
    if not column_index:
        return None
    cell: Cell = worksheet.cell(row=row_index, column=column_index)
    if cell.hyperlink is not None and cell.hyperlink.target:
        return cell.hyperlink.target
    return cell.value


def _text(value: object) -> str | None:
    if value is None:
        return None
    compact = re.sub(r"[\t ]+", " ", str(value)).strip()
    return compact or None


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _non_negative_int(value: object) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return number if number >= 0 else None


def _contact_id(inn: str, full_name: str, role: ContactRole):
    identity = f"kontur:{inn}:{full_name.casefold()}:{role.value}"
    return uuid5(CONTACT_NAMESPACE, identity)


def _issue(row_index: int, code: str, message: str) -> ImportIssue:
    return ImportIssue(
        row_index=row_index,
        code=code,
        message=message,
        severity=ImportIssueSeverity.ERROR,
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
