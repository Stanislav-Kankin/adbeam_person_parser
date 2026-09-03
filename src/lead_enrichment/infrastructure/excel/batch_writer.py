from __future__ import annotations

import os
import tempfile
from collections.abc import Iterable
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lead_enrichment.infrastructure.excel.writer import sanitize_excel_text
from lead_enrichment.models import (
    BatchEnrichmentResult,
    ChannelType,
    ContactChannel,
    EnrichedLeadResult,
    SourceReference,
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
REQUIRED_SHEETS = {
    "Сводка",
    "Компании",
    "Контакты",
    "Источники",
    "Журнал",
    "Ручная очередь",
    "Исходные данные",
}


def export_batch_enrichment_workbook(
    result: BatchEnrichmentResult,
    output_file_path: Path,
) -> Path:
    target = Path(output_file_path)
    if target.suffix.casefold() != ".xlsx":
        raise ValueError("Итоговый файл должен иметь расширение .xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    try:
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Сводка")
        companies = workbook.create_sheet("Компании")
        contacts = workbook.create_sheet("Контакты")
        sources = workbook.create_sheet("Источники")
        journal = workbook.create_sheet("Журнал")
        manual = workbook.create_sheet("Ручная очередь")
        source_data = workbook.create_sheet("Исходные данные")

        _write_companies(companies, result.leads)
        _write_contacts(contacts, result.leads)
        _write_sources(sources, result.leads)
        _write_journal(journal, result.leads)
        _write_manual_queue(manual, result.leads)
        _write_source_data(source_data, result.leads)
        _write_summary(summary, result, companies.max_row, journal.max_row)

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        _atomic_save(workbook, target)
    finally:
        workbook.close()
    return target


def _write_summary(sheet, result: BatchEnrichmentResult, company_rows: int, journal_rows: int) -> None:
    sheet.append(["AdBeam Person Parser — результат обработки", None])
    sheet.append(["Run ID", _safe_value(result.run_id)])
    sheet.append(["Целевые роли", "; ".join(role.value for role in result.target_roles)])
    sheet.append([None, None])
    sheet.append(["Показатель", "Значение"])
    company_end = max(company_rows, 2)
    journal_end = max(journal_rows, 2)
    sheet.append(["Всего компаний", f"=COUNTA('Компании'!A2:A{company_end})"])
    sheet.append(["RESOLVED", f'=COUNTIF(\'Компании\'!L2:L{company_end},"RESOLVED")'])
    sheet.append(["PARTIAL", f'=COUNTIF(\'Компании\'!L2:L{company_end},"PARTIAL")'])
    sheet.append([
        "MANUAL_REQUIRED",
        f'=COUNTIF(\'Компании\'!L2:L{company_end},"MANUAL_REQUIRED")',
    ])
    sheet.append([
        "Checkpoint hits",
        f"=COUNTIF('Журнал'!G2:G{journal_end},TRUE)",
    ])
    sheet.append([
        "Выполнений источников",
        f"=COUNTIF('Журнал'!G2:G{journal_end},FALSE)",
    ])
    sheet.merge_cells("A1:B1")
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = Font(bold=True, size=14, color="1F1F1F")
    sheet["A5"].fill = HEADER_FILL
    sheet["B5"].fill = HEADER_FILL
    sheet["A5"].font = HEADER_FONT
    sheet["B5"].font = HEADER_FONT
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 42
    sheet.sheet_view.showGridLines = False


def _write_companies(sheet, leads: list[EnrichedLeadResult]) -> None:
    headers = [
        "Company key",
        "Наименование",
        "ИНН",
        "КПП",
        "ОГРН / ОГРНИП",
        "Тип",
        "Регион",
        "Сайт",
        "Сегмент",
        "Основной вид деятельности",
        "Статус организации",
        "Coverage status",
        "Причина coverage",
        "Email компании",
        "Телефоны компании",
        "Соцсети компании",
        "ЛПР найдено",
        "Недостающие роли",
        "Identity status",
        "Метод сопоставления",
        "Уверенность identity",
    ]
    sheet.append(headers)
    for item in leads:
        company = item.company
        sheet.append([
            _safe_value(company.company_key),
            _safe_value(_company_name(company)),
            company.inn,
            company.kpp,
            company.ogrn,
            company.entity_type.value,
            _safe_value(company.region),
            _safe_value(company.website),
            _safe_value(company.segment_name),
            _safe_value(company.primary_activity),
            _safe_value(company.operating_status),
            item.coverage.status.value,
            _safe_value(item.coverage.reason_message),
            _safe_value(_channel_values(company.company_channels, ChannelType.EMAIL)),
            _safe_value(_channel_values(company.company_channels, ChannelType.PHONE)),
            _safe_value(_channel_values(company.company_channels, ChannelType.SOCIAL)),
            len(company.initial_people),
            "; ".join(role.value for role in item.coverage.missing_roles),
            item.lead.identity.status.value,
            item.lead.identity.method.value,
            item.lead.identity.confidence_score,
        ])
        _set_hyperlink(sheet.cell(row=sheet.max_row, column=8), company.website)
    _format_table(sheet, widths=[34, 42, 15, 12, 18, 22, 24, 36, 28, 46, 22, 18, 46, 38, 34, 36, 12, 28, 18, 20, 16])


def _write_contacts(sheet, leads: list[EnrichedLeadResult]) -> None:
    headers = [
        "Company key",
        "Наименование",
        "ИНН",
        "ФИО",
        "Должность",
        "Роль",
        "Уверенность",
        "Персональный email",
        "Персональный телефон",
        "Персональные соцсети",
        "Источники",
        "Ссылки подтверждения",
    ]
    sheet.append(headers)
    for item in leads:
        for person in item.company.initial_people:
            refs = _unique_refs(
                [*person.source_refs, *(ref for channel in person.channels for ref in channel.source_refs)]
            )
            sheet.append([
                _safe_value(item.company.company_key),
                _safe_value(_company_name(item.company)),
                item.company.inn,
                _safe_value(person.full_name),
                _safe_value(person.job_title),
                person.normalized_role.value,
                person.confidence_score,
                _safe_value(_channel_values(person.channels, ChannelType.EMAIL)),
                _safe_value(_channel_values(person.channels, ChannelType.PHONE)),
                _safe_value(_channel_values(person.channels, ChannelType.SOCIAL)),
                _safe_value("; ".join(dict.fromkeys(ref.source_name for ref in refs))),
                _safe_value("\n".join(ref.url for ref in refs if ref.url)),
            ])
    _format_table(sheet, widths=[34, 32, 15, 32, 34, 18, 14, 36, 30, 36, 34, 55])


def _write_sources(sheet, leads: list[EnrichedLeadResult]) -> None:
    headers = [
        "Company key",
        "Источник",
        "Тип",
        "Locator",
        "Дата сбора",
        "URL",
        "Надёжность",
    ]
    sheet.append(headers)
    for item in leads:
        for ref in _company_refs(item):
            sheet.append([
                _safe_value(item.company.company_key),
                _safe_value(ref.source_name),
                ref.source_kind.value,
                _safe_value(ref.locator),
                _excel_datetime(ref.collected_at),
                _safe_value(ref.url),
                ref.reliability,
            ])
            _set_hyperlink(sheet.cell(row=sheet.max_row, column=6), ref.url)
    _format_table(sheet, widths=[34, 32, 24, 30, 22, 55, 14])
    for cell in sheet["E"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm:ss"


def _write_journal(sheet, leads: list[EnrichedLeadResult]) -> None:
    headers = [
        "Company key",
        "Источник",
        "Версия",
        "Результат",
        "Код причины",
        "Описание",
        "Из checkpoint",
        "Длительность, мс",
    ]
    sheet.append(headers)
    for item in leads:
        for step in item.steps:
            sheet.append([
                _safe_value(step.company_key),
                step.source_id,
                step.source_version,
                step.outcome.value,
                step.reason_code,
                _safe_value(step.reason_message),
                step.from_checkpoint,
                step.duration_ms,
            ])
    _format_table(sheet, widths=[34, 24, 12, 16, 30, 54, 16, 18])


def _write_manual_queue(sheet, leads: list[EnrichedLeadResult]) -> None:
    headers = [
        "Наименование",
        "ИНН",
        "Причина",
        "Недостающие роли",
        "Поисковая ссылка",
    ]
    sheet.append(headers)
    for item in leads:
        for url in item.manual_search_urls:
            sheet.append([
                _safe_value(_company_name(item.company)),
                item.company.inn,
                _safe_value(item.coverage.reason_message),
                "; ".join(role.value for role in item.coverage.missing_roles),
                url,
            ])
            _set_hyperlink(sheet.cell(row=sheet.max_row, column=5), url)
    _format_table(sheet, widths=[42, 15, 50, 28, 70])


def _write_source_data(sheet, leads: list[EnrichedLeadResult]) -> None:
    sheet.append([
        "Company key",
        "Строка источника",
        "Наименование",
        "ИНН",
        "КПП",
        "ОГРН / ОГРНИП",
        "Тип",
        "Регион",
        "Адрес",
        "Статус организации",
        "Реестр МСП",
        "Карточка в Фокусе",
        "Сайт",
        "Основной вид деятельности",
        "Количество сотрудников",
        "Количество филиалов",
        "Источник",
        "Сегмент",
        "Email компании",
        "Телефоны компании",
    ])
    for item in leads:
        company = item.lead.kontur_company or item.company
        sheet.append([
            _safe_value(company.company_key),
            _safe_value(company.input_row_id),
            _safe_value(_company_name(company)),
            company.inn,
            company.kpp,
            company.ogrn,
            company.entity_type.value,
            _safe_value(company.region),
            _safe_value(company.address),
            _safe_value(company.operating_status),
            _safe_value(company.msp_category),
            _safe_value(company.focus_url),
            _safe_value(company.website),
            _safe_value(company.primary_activity),
            company.employee_count,
            company.branch_count,
            _safe_value(company.source_label),
            _safe_value(company.segment_name),
            _safe_value(_channel_values(company.company_channels, ChannelType.EMAIL)),
            _safe_value(_channel_values(company.company_channels, ChannelType.PHONE)),
        ])
        _set_hyperlink(sheet.cell(row=sheet.max_row, column=12), company.focus_url)
        _set_hyperlink(sheet.cell(row=sheet.max_row, column=13), company.website)
    _format_table(
        sheet,
        widths=[34, 22, 42, 15, 12, 18, 22, 24, 48, 22, 20, 40, 36, 50, 18, 18, 26, 28, 38, 34],
    )


def _company_name(company) -> str:
    return company.brand_name or company.legal_name or company.company_key


def _company_refs(item: EnrichedLeadResult) -> list[SourceReference]:
    refs: list[SourceReference] = list(item.company.source_refs)
    for channel in item.company.company_channels:
        refs.extend(channel.source_refs)
    for person in item.company.initial_people:
        refs.extend(person.source_refs)
        for channel in person.channels:
            refs.extend(channel.source_refs)
    return _unique_refs(refs)


def _unique_refs(refs: Iterable[SourceReference]) -> list[SourceReference]:
    result: dict[tuple[str, str, str | None], SourceReference] = {}
    for ref in refs:
        result.setdefault((ref.source_id, ref.locator, ref.url), ref)
    return list(result.values())


def _channel_values(channels: Iterable[ContactChannel], channel_type: ChannelType) -> str | None:
    values = list(
        dict.fromkeys(
            channel.value
            for channel in channels
            if channel.channel_type == channel_type
        )
    )
    return "; ".join(values) or None


def _format_table(sheet, *, widths: list[int]) -> None:
    if sheet.max_column:
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"
        for index in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(index)].width = min(
                widths[index - 1] if index <= len(widths) else 24,
                70,
            )
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.sheet_view.showGridLines = False


def _atomic_save(workbook: Workbook, target: Path) -> None:
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{target.stem}.",
        suffix=".tmp.xlsx",
        dir=target.parent,
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        try:
            workbook.save(temp_path)
        finally:
            workbook.close()
        _validate_export(temp_path)
        os.replace(temp_path, target)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _validate_export(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        missing = REQUIRED_SHEETS - set(workbook.sheetnames)
        if missing:
            raise ValueError("В итоговой книге отсутствуют обязательные листы")
        if workbook["Компании"]["A1"].value != "Company key":
            raise ValueError("В итоговой книге повреждён лист компаний")
        if workbook["Контакты"]["D1"].value != "ФИО":
            raise ValueError("В итоговой книге повреждён лист контактов")
    finally:
        workbook.close()


def _safe_value(value):
    if isinstance(value, datetime):
        return _excel_datetime(value)
    if value is None or isinstance(value, (int, float, bool, date)):
        return value
    return sanitize_excel_text(str(value))


def _excel_datetime(value: datetime) -> datetime:
    if value.tzinfo is None or value.utcoffset() is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _set_hyperlink(cell, value: str | None) -> None:
    if not value or not _is_safe_http_url(value):
        return
    cell.hyperlink = value
    cell.style = "Hyperlink"


def _is_safe_http_url(value: str) -> bool:
    try:
        parsed = urlsplit(value)
    except ValueError:
        return False
    return bool(
        parsed.scheme.casefold() in {"http", "https"}
        and parsed.hostname
        and not parsed.username
        and not parsed.password
    )
