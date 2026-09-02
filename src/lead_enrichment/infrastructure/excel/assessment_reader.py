from __future__ import annotations

import hashlib
import re
import warnings
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from openpyxl import load_workbook

from lead_enrichment.engine.normalization import (
    normalize_http_url,
    normalize_person_name,
    split_phones,
)
from lead_enrichment.models import (
    AssessmentCompany,
    AssessmentContact,
    AssessmentImportResult,
    AssessmentImportSummary,
    AssessmentScores,
    ChannelScope,
    ChannelType,
    ContactChannel,
    ContactRole,
    ImportIssue,
    ImportIssueSeverity,
    LeadPriority,
    SourceKind,
    SourceReference,
)

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
DEFAULT_MAIN_SHEET = "Ростовская область"
DEFAULT_LPR_SHEET = "ЛПР 2026 verified"

MAIN_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "brand_name": ("Застройщик / бренд",),
    "region": ("Регион",),
    "assessment_date": ("Дата ассесмента",),
    "source_rating_date": ("Дата рейтинга источника",),
    "erz_id": ("ID ЕРЗ",),
    "erz_url": ("Ссылка на карточку ЕРЗ",),
    "website": ("Сайт компании / проекта",),
    "projects": ("Активные ЖК по ЕРЗ",),
    "project_cities": ("Города активных ЖК по ЕРЗ",),
    "address": ("Адрес по ЕРЗ",),
    "sales_phones": ("Телефон отдела продаж по ЕРЗ",),
    "vk": ("VK по ЕРЗ",),
    "scale_type": ("Тип масштаба застройщика",),
    "stage1_status": ("Итог Stage 1",),
    "indigo_match_status": ("Статус совпадения с Indigo",),
    "matched_indigo_group": ("Совпавшая группа Indigo",),
    "matched_indigo_alias": ("Совпавший алиас / проект",),
    "indigo_match_type": ("Тип совпадения Indigo",),
    "workflow_status": ("Статус текущей проработки",),
    "owner": ("Ответственный",),
    "last_touch_date": ("Дата последнего касания",),
    "commercial_score": ("Баллы: коммерческий потенциал",),
    "marketing_score": ("Баллы: маркетинговая потребность",),
    "digital_score": ("Баллы: digital-возможность",),
    "outreach_score": ("Баллы: outreach-доступность",),
    "total_score": ("Итоговый балл",),
    "lead_priority": ("TIR",),
    "cap_rule": ("Примененное cap-правило",),
    "outreach_hook": ("Зацепка для первого касания",),
    "next_action": ("Следующее действие",),
    "missing_data": ("Недостающие данные",),
    "source_urls": ("Ссылки на источники",),
    "comments": ("Комментарии",),
    "lpr_source_urls": ("Источники ЛПР",),
}

LPR_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "main_row": ("Строка",),
    "company": ("Компания",),
    "projects": ("Проекты / ЖК",),
    "primary_name": ("Основной актуальный ЛПР",),
    "primary_role": ("Роль",),
    "alternative": ("Альтернативный / стратегический ЛПР",),
    "confirmation": ("Источник / подтверждение",),
    "verification_status": ("Статус актуальности",),
    "outreach_recommendation": ("Рекомендация для outreach",),
    "comment": ("Комментарий",),
    "links": ("Ссылки",),
}


class AssessmentWorkbookError(ValueError):
    """The workbook cannot be interpreted as a supported client assessment."""


class MissingAssessmentColumnsError(AssessmentWorkbookError):
    def __init__(self, sheet_name: str, missing_columns: list[str]) -> None:
        self.sheet_name = sheet_name
        self.missing_columns = missing_columns
        super().__init__(
            f"На листе {sheet_name!r} отсутствуют обязательные колонки: "
            f"{', '.join(missing_columns)}"
        )


def read_assessment_workbook(
    file_path: Path,
    *,
    main_sheet_name: str = DEFAULT_MAIN_SHEET,
    lpr_sheet_name: str = DEFAULT_LPR_SHEET,
    collected_at: datetime | None = None,
) -> AssessmentImportResult:
    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError("Файл клиентского assessment не найден")
    if path.suffix.casefold() not in SUPPORTED_SUFFIXES:
        raise AssessmentWorkbookError("Поддерживаются только файлы .xlsx и .xlsm")

    collected_at = collected_at or datetime.now(timezone.utc)
    if collected_at.tzinfo is None or collected_at.utcoffset() is None:
        raise ValueError("collected_at must be timezone-aware")

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style.*")
        workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)

    try:
        main_sheet = _required_sheet(workbook, main_sheet_name)
        lpr_sheet = _required_sheet(workbook, lpr_sheet_name)
        main_columns = _build_column_map(main_sheet, MAIN_COLUMN_ALIASES)
        lpr_columns = _build_column_map(lpr_sheet, LPR_COLUMN_ALIASES)
        _require_columns(main_sheet_name, main_columns, {"brand_name"}, MAIN_COLUMN_ALIASES)
        _require_columns(
            lpr_sheet_name,
            lpr_columns,
            {"main_row", "company", "primary_name", "primary_role"},
            LPR_COLUMN_ALIASES,
        )

        issues: list[ImportIssue] = []
        contacts_by_main_row = _read_lpr_contacts(
            lpr_sheet,
            lpr_columns,
            main_sheet,
            main_columns,
            collected_at,
            issues,
        )
        companies: list[AssessmentCompany] = []
        seen_brand_keys: set[str] = set()
        seen_company_keys: set[str] = set()
        rows_with_website = 0
        rows_with_sales_phones = 0
        primary_contacts = 0
        alternative_contacts = 0
        indigo_matches = 0
        tier_counts: Counter[str] = Counter()

        for row_index in range(2, main_sheet.max_row + 1):
            if _row_is_blank(main_sheet, row_index):
                continue
            brand_name = _text(_value(main_sheet, row_index, main_columns, "brand_name"))
            if not brand_name:
                issues.append(_issue(row_index, "MISSING_BRAND_NAME", "Не заполнено название бренда"))
                continue

            region = _text(_value(main_sheet, row_index, main_columns, "region"))
            brand_key = _normalize_name(brand_name)
            if brand_key in seen_brand_keys:
                issues.append(
                    _issue(
                        row_index,
                        "DUPLICATE_BRAND_NAME",
                        "Нормализованное название бренда уже встречалось; строки не объединены",
                        severity=ImportIssueSeverity.WARNING,
                    )
                )
            seen_brand_keys.add(brand_key)

            website = normalize_http_url(_value(main_sheet, row_index, main_columns, "website"))
            erz_id = _identifier_text(_value(main_sheet, row_index, main_columns, "erz_id"))
            company_key = _company_key(erz_id, website, brand_name, region)
            if company_key in seen_company_keys:
                issues.append(
                    _issue(
                        row_index,
                        "DUPLICATE_COMPANY_KEY",
                        "Временный ключ компании уже встречался; строки не объединены",
                        severity=ImportIssueSeverity.WARNING,
                    )
                )
            seen_company_keys.add(company_key)

            source_urls = _unique(
                _split_urls(_value(main_sheet, row_index, main_columns, "source_urls"))
                + _split_urls(_value(main_sheet, row_index, main_columns, "lpr_source_urls"))
            )
            erz_url = normalize_http_url(_value(main_sheet, row_index, main_columns, "erz_url"))
            source_ref = SourceReference(
                source_id="client_assessment",
                source_kind=SourceKind.CLIENT_ASSESSMENT,
                source_name="Клиентский assessment",
                locator=f"{main_sheet.title}!{row_index}",
                collected_at=collected_at,
                url=erz_url,
                reliability=75,
            )
            phones = split_phones(
                _value(main_sheet, row_index, main_columns, "sales_phones"),
                allow_ten_digit=True,
            )
            channels = [
                ContactChannel(
                    channel_type=ChannelType.PHONE,
                    value=phone,
                    scope=ChannelScope.COMPANY,
                    source_refs=[source_ref],
                )
                for phone in phones
            ]
            vk_url = normalize_http_url(_value(main_sheet, row_index, main_columns, "vk"))
            if vk_url:
                channels.append(
                    ContactChannel(
                        channel_type=ChannelType.SOCIAL,
                        value=vk_url,
                        scope=ChannelScope.COMPANY,
                        source_refs=[source_ref],
                    )
                )

            contacts = contacts_by_main_row.get(row_index, [])
            priority = _lead_priority(_value(main_sheet, row_index, main_columns, "lead_priority"))
            tier_counts[priority.value] += 1
            match_status = _text(
                _value(main_sheet, row_index, main_columns, "indigo_match_status")
            )
            rows_with_website += int(website is not None)
            rows_with_sales_phones += int(bool(phones))
            primary_contacts += sum(contact.is_primary for contact in contacts)
            alternative_contacts += sum(not contact.is_primary for contact in contacts)
            indigo_matches += int(bool(match_status and match_status.casefold() != "no_match"))

            companies.append(
                AssessmentCompany(
                    company_key=company_key,
                    input_row_id=f"{main_sheet.title}:{row_index}",
                    assessment_row=row_index,
                    brand_name=brand_name,
                    region=region,
                    assessment_date=_date_value(
                        _value(main_sheet, row_index, main_columns, "assessment_date")
                    ),
                    source_rating_date=_date_value(
                        _value(main_sheet, row_index, main_columns, "source_rating_date")
                    ),
                    erz_id=erz_id,
                    erz_url=erz_url,
                    website=website,
                    address=_text(_value(main_sheet, row_index, main_columns, "address")),
                    projects=_split_list(_value(main_sheet, row_index, main_columns, "projects")),
                    project_cities=_split_list(
                        _value(main_sheet, row_index, main_columns, "project_cities")
                    ),
                    company_channels=channels,
                    contacts=contacts,
                    scale_type=_text(_value(main_sheet, row_index, main_columns, "scale_type")),
                    stage1_status=_text(
                        _value(main_sheet, row_index, main_columns, "stage1_status")
                    ),
                    lead_priority=priority,
                    cap_rule=_text(_value(main_sheet, row_index, main_columns, "cap_rule")),
                    scores=AssessmentScores(
                        commercial_potential=_non_negative_int(
                            _value(main_sheet, row_index, main_columns, "commercial_score")
                        ),
                        marketing_need=_non_negative_int(
                            _value(main_sheet, row_index, main_columns, "marketing_score")
                        ),
                        digital_opportunity=_non_negative_int(
                            _value(main_sheet, row_index, main_columns, "digital_score")
                        ),
                        outreach_accessibility=_non_negative_int(
                            _value(main_sheet, row_index, main_columns, "outreach_score")
                        ),
                        total=_non_negative_int(
                            _value(main_sheet, row_index, main_columns, "total_score")
                        ),
                    ),
                    indigo_match_status=match_status,
                    matched_indigo_group=_text(
                        _value(main_sheet, row_index, main_columns, "matched_indigo_group")
                    ),
                    matched_indigo_alias=_text(
                        _value(main_sheet, row_index, main_columns, "matched_indigo_alias")
                    ),
                    indigo_match_type=_text(
                        _value(main_sheet, row_index, main_columns, "indigo_match_type")
                    ),
                    workflow_status=_text(
                        _value(main_sheet, row_index, main_columns, "workflow_status")
                    ),
                    owner=_text(_value(main_sheet, row_index, main_columns, "owner")),
                    last_touch_date=_date_value(
                        _value(main_sheet, row_index, main_columns, "last_touch_date")
                    ),
                    outreach_hook=_text(
                        _value(main_sheet, row_index, main_columns, "outreach_hook")
                    ),
                    next_action=_text(
                        _value(main_sheet, row_index, main_columns, "next_action")
                    ),
                    missing_data=_text(
                        _value(main_sheet, row_index, main_columns, "missing_data")
                    ),
                    source_urls=source_urls,
                    comments=_text(_value(main_sheet, row_index, main_columns, "comments")),
                    source_fields=_source_fields(main_sheet, row_index),
                    source_refs=[source_ref],
                )
            )

        total_rows = sum(
            not _row_is_blank(main_sheet, row_index)
            for row_index in range(2, main_sheet.max_row + 1)
        )
        return AssessmentImportResult(
            summary=AssessmentImportSummary(
                source_file_name=path.name,
                source_sha256=_sha256(path),
                main_sheet_name=main_sheet.title,
                lpr_sheet_name=lpr_sheet.title,
                total_rows=total_rows,
                imported_rows=len(companies),
                skipped_rows=total_rows - len(companies),
                rows_with_website=rows_with_website,
                rows_with_sales_phones=rows_with_sales_phones,
                primary_contacts=primary_contacts,
                alternative_contacts=alternative_contacts,
                indigo_matches=indigo_matches,
                tier_counts=dict(tier_counts),
                issues=issues,
            ),
            companies=companies,
        )
    finally:
        workbook.close()


def _read_lpr_contacts(
    worksheet,
    column_map: dict[str, int],
    main_sheet,
    main_column_map: dict[str, int],
    collected_at: datetime,
    issues: list[ImportIssue],
) -> dict[int, list[AssessmentContact]]:
    result: dict[int, list[AssessmentContact]] = {}
    for row_index in range(2, worksheet.max_row + 1):
        if _row_is_blank(worksheet, row_index):
            continue
        main_row = _positive_int(_value(worksheet, row_index, column_map, "main_row"))
        if main_row is None or main_row < 2 or main_row > main_sheet.max_row:
            issues.append(
                _issue(
                    row_index,
                    "INVALID_LPR_ROW_REFERENCE",
                    "Ссылка на строку основного листа отсутствует или выходит за его границы",
                )
            )
            continue
        if main_row in result:
            issues.append(
                _issue(
                    row_index,
                    "DUPLICATE_LPR_ROW_REFERENCE",
                    "Строка основного листа уже связана с другой строкой ЛПР",
                )
            )
            continue

        lpr_company = _text(_value(worksheet, row_index, column_map, "company"))
        main_company = _text(_value(main_sheet, main_row, main_column_map, "brand_name"))
        if not lpr_company or not main_company or _normalize_name(lpr_company) != _normalize_name(
            main_company
        ):
            issues.append(
                _issue(
                    row_index,
                    "LPR_COMPANY_MISMATCH",
                    "Компания в строке ЛПР не совпадает с указанной строкой основного листа",
                )
            )
            continue

        evidence_urls = _unique(
            _split_urls(_value(worksheet, row_index, column_map, "confirmation"))
            + _split_urls(_value(worksheet, row_index, column_map, "links"))
        )
        source_refs = _contact_source_refs(
            worksheet.title,
            row_index,
            evidence_urls,
            collected_at,
        )
        verification_status = _text(
            _value(worksheet, row_index, column_map, "verification_status")
        )
        outreach_recommendation = _text(
            _value(worksheet, row_index, column_map, "outreach_recommendation")
        )
        comment = _text(_value(worksheet, row_index, column_map, "comment"))
        contacts: list[AssessmentContact] = []
        primary_name = normalize_person_name(
            _value(worksheet, row_index, column_map, "primary_name")
        )
        primary_title = _text(_value(worksheet, row_index, column_map, "primary_role"))
        if primary_name:
            contacts.append(
                AssessmentContact(
                    full_name=primary_name,
                    job_title=primary_title,
                    normalized_role=_contact_role(primary_title),
                    is_primary=True,
                    verification_status=verification_status,
                    outreach_recommendation=outreach_recommendation,
                    comment=comment,
                    evidence_urls=evidence_urls,
                    source_refs=source_refs,
                )
            )
        else:
            issues.append(
                _issue(
                    row_index,
                    "MISSING_PRIMARY_LPR",
                    "В связанной строке отсутствует основной ЛПР",
                    severity=ImportIssueSeverity.WARNING,
                )
            )

        for full_name, job_title in _parse_alternative_contacts(
            _value(worksheet, row_index, column_map, "alternative")
        ):
            contacts.append(
                AssessmentContact(
                    full_name=full_name,
                    job_title=job_title,
                    normalized_role=_contact_role(job_title),
                    is_primary=False,
                    verification_status=verification_status,
                    outreach_recommendation=outreach_recommendation,
                    comment=comment,
                    evidence_urls=evidence_urls,
                    source_refs=source_refs,
                )
            )
        result[main_row] = contacts
    return result


def _parse_alternative_contacts(value: object) -> list[tuple[str, str | None]]:
    text = _text(value)
    if not text:
        return []
    contacts: list[tuple[str, str | None]] = []
    for part in (item.strip() for item in text.split(";") if item.strip()):
        if " — " not in part:
            if contacts:
                full_name, title = contacts[-1]
                contacts[-1] = (full_name, "; ".join(filter(None, [title, part])))
            continue
        name, title = part.split(" — ", 1)
        normalized_name = normalize_person_name(name)
        if normalized_name:
            contacts.append((normalized_name, _text(title)))
    return contacts


def _contact_source_refs(
    sheet_name: str,
    row_index: int,
    urls: list[str],
    collected_at: datetime,
) -> list[SourceReference]:
    if not urls:
        return [
            SourceReference(
                source_id="client_assessment_lpr",
                source_kind=SourceKind.CLIENT_ASSESSMENT,
                source_name="Клиентский assessment — ЛПР",
                locator=f"{sheet_name}!{row_index}",
                collected_at=collected_at,
                reliability=70,
            )
        ]
    return [
        SourceReference(
            source_id="client_assessment_lpr",
            source_kind=SourceKind.CLIENT_ASSESSMENT,
            source_name="Клиентский assessment — ЛПР",
            locator=f"{sheet_name}!{row_index}",
            collected_at=collected_at,
            url=url,
            reliability=70,
        )
        for url in urls
    ]


def _required_sheet(workbook: Any, name: str):
    if name not in workbook.sheetnames:
        raise AssessmentWorkbookError(f"В книге отсутствует обязательный лист {name!r}")
    return workbook[name]


def _build_column_map(worksheet, aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
    observed: dict[str, int] = {}
    for column_index, cell in enumerate(worksheet[1], start=1):
        normalized = _normalize_header(cell.value)
        if not normalized:
            continue
        if normalized in observed:
            raise AssessmentWorkbookError(
                f"На листе {worksheet.title!r} есть дублирующиеся заголовки"
            )
        observed[normalized] = column_index
    return {
        field_name: observed[_normalize_header(alias)]
        for field_name, field_aliases in aliases.items()
        for alias in field_aliases
        if _normalize_header(alias) in observed
    }


def _require_columns(
    sheet_name: str,
    column_map: dict[str, int],
    required: set[str],
    aliases: dict[str, tuple[str, ...]],
) -> None:
    missing = sorted(required - column_map.keys())
    if missing:
        raise MissingAssessmentColumnsError(
            sheet_name,
            [aliases[field_name][0] for field_name in missing],
        )


def _source_fields(worksheet, row_index: int) -> dict[str, str | int | float | date | None]:
    result: dict[str, str | int | float | date | None] = {}
    for cell in worksheet[row_index]:
        header = _text(worksheet.cell(row=1, column=cell.column).value)
        if not header:
            continue
        value = cell.value
        if isinstance(value, datetime):
            value = value.date()
        if value is None or isinstance(value, (str, int, float, date)):
            result[header] = value
        else:
            result[header] = str(value)
    return result


def _company_key(
    erz_id: str | None,
    website: str | None,
    brand_name: str,
    region: str | None,
) -> str:
    if erz_id:
        normalized_id = re.sub(r"\s+", "", erz_id).casefold()
        return f"assessment:erz:{normalized_id}"
    if website:
        hostname = (urlsplit(website).hostname or "").casefold()
        if hostname.startswith("www."):
            hostname = hostname[4:]
        if hostname:
            return f"assessment:domain:{hostname}"
    identity = f"{_normalize_name(brand_name)}|{_normalize_name(region or '')}"
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:24]
    return f"assessment:brand:{digest}"


def _lead_priority(value: object) -> LeadPriority:
    normalized = _text(value)
    try:
        return LeadPriority(normalized)
    except ValueError:
        return LeadPriority.UNKNOWN


def _contact_role(value: object) -> ContactRole:
    normalized = (_text(value) or "").casefold()
    if "маркет" in normalized:
        return ContactRole.MARKETING
    if "продаж" in normalized or "коммерч" in normalized:
        return ContactRole.SALES
    if "закуп" in normalized or "снабж" in normalized:
        return ContactRole.PROCUREMENT
    if any(term in normalized for term in ("собствен", "учред", "основател")):
        return ContactRole.OWNER
    if any(term in normalized for term in ("директор", "руководител", "управляющ")):
        return ContactRole.LEADER
    return ContactRole.UNKNOWN


def _split_urls(value: object) -> list[str]:
    if value is None:
        return []
    urls: list[str] = []
    for match in re.findall(r"https?://[^\s;,|]+", str(value), flags=re.IGNORECASE):
        normalized = normalize_http_url(match.rstrip(".)]}>"))
        if normalized:
            urls.append(normalized)
    return _unique(urls)


def _split_list(value: object) -> list[str]:
    if value is None:
        return []
    return _unique(
        part.strip()
        for part in re.split(r"[\r\n;|]+", str(value))
        if part.strip()
    )


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    raw = str(value).strip()
    for format_string in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, format_string).date()
        except ValueError:
            continue
    return None


def _identifier_text(value: object) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _positive_int(value: object) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return number if number > 0 else None


def _non_negative_int(value: object) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return number if number >= 0 else None


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    normalized = re.sub(r"\s+", " ", str(value)).strip().casefold()
    return re.sub(r"\s*/\s*", "/", normalized)


def _normalize_name(value: str) -> str:
    return re.sub(r"[^0-9a-zа-яё]+", "", value.casefold())


def _row_is_blank(worksheet, row_index: int) -> bool:
    return not any(
        cell.value is not None and str(cell.value).strip() for cell in worksheet[row_index]
    )


def _value(worksheet, row_index: int, column_map: dict[str, int], field_name: str):
    column_index = column_map.get(field_name)
    return worksheet.cell(row=row_index, column=column_index).value if column_index else None


def _text(value: object) -> str | None:
    if value is None:
        return None
    compact = re.sub(r"[\t ]+", " ", str(value)).strip()
    return compact or None


def _issue(
    row_index: int,
    code: str,
    message: str,
    *,
    severity: ImportIssueSeverity = ImportIssueSeverity.ERROR,
) -> ImportIssue:
    return ImportIssue(
        row_index=row_index,
        code=code,
        message=message,
        severity=severity,
    )


def _unique(values) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value:
            continue
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
