from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from lead_enrichment.infrastructure.excel import export_enrichment_workbook, sanitize_excel_text
from lead_enrichment.infrastructure.excel import writer as writer_module
from lead_enrichment.models import (
    ChannelScope,
    ChannelType,
    CompanyInput,
    ContactChannel,
    ContactRole,
    EntityType,
    KonturImportResult,
    KonturImportSummary,
    PersonContact,
    SourceKind,
    SourceReference,
)


def _import_result(*, dangerous_name: bool = False) -> KonturImportResult:
    collected_at = datetime(2026, 9, 2, tzinfo=timezone.utc)
    source_ref = SourceReference(
        source_id="kontur_import",
        source_kind=SourceKind.KONTUR_EXPORT,
        source_name="Контур.Продажи",
        locator="Контрагенты!2",
        collected_at=collected_at,
        reliability=90,
    )
    company = CompanyInput(
        input_row_id="Контрагенты:2",
        legal_name="=2+2" if dangerous_name else 'ООО "Тест"',
        inn="1234567894",
        entity_type=EntityType.LEGAL_ENTITY,
        website="https://example.com/",
        company_channels=[
            ContactChannel(
                channel_type=ChannelType.EMAIL,
                value="hello@example.com",
                scope=ChannelScope.COMPANY,
                source_refs=[source_ref],
            )
        ],
        initial_people=[
            PersonContact(
                contact_id="51f27b8e-66f3-4d25-9ca0-5d811ad1a89c",
                full_name="Иванов Иван Иванович",
                job_title="Генеральный директор",
                normalized_role=ContactRole.LEADER,
                confidence_score=90,
                source_refs=[source_ref],
            )
        ],
        source_refs=[source_ref],
    )
    summary = KonturImportSummary(
        source_file_name="synthetic.xlsx",
        source_sha256="a" * 64,
        sheet_name="Контрагенты",
        total_rows=1,
        imported_rows=1,
        skipped_rows=0,
        blank_rows=0,
        duplicate_inn_rows=0,
        legal_entities=1,
        individual_entrepreneurs=0,
        rows_with_initial_person=1,
        rows_with_phones=0,
        rows_with_emails=1,
        rows_with_website=1,
    )
    return KonturImportResult(summary=summary, companies=[company])


def test_sanitize_excel_text_removes_xml_chars_limits_length_and_blocks_formula() -> None:
    assert sanitize_excel_text("abc\x00def") == "abcdef"
    assert sanitize_excel_text(" =1+1") == "' =1+1"
    assert len(sanitize_excel_text("x" * 40_000)) == 32_767


def test_export_creates_valid_workbook_with_preserved_hyperlink(tmp_path: Path) -> None:
    output = tmp_path / "result.xlsx"

    export_enrichment_workbook(_import_result(dangerous_name=True), output)

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == ["Сводка", "Компании", "Контакты", "Источники", "Журнал"]
        companies = workbook["Компании"]
        assert companies["B2"].value == "'=2+2"
        assert companies["H2"].hyperlink.target == "https://example.com/"
        assert workbook["Контакты"]["F2"].value == "hello@example.com"
        assert workbook["Сводка"]["B7"].value == 1
    finally:
        workbook.close()


def test_atomic_export_keeps_existing_target_when_validation_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "result.xlsx"
    sentinel = Workbook()
    sentinel.active["A1"] = "old"
    sentinel.save(output)
    sentinel.close()
    original_bytes = output.read_bytes()

    def fail_validation(_path: Path) -> None:
        raise ValueError("synthetic validation failure")

    monkeypatch.setattr(writer_module, "_validate_export_workbook", fail_validation)

    with pytest.raises(ValueError, match="synthetic validation failure"):
        export_enrichment_workbook(_import_result(), output)

    assert output.read_bytes() == original_bytes
    assert list(tmp_path.glob(".result.*.tmp.xlsx")) == []
