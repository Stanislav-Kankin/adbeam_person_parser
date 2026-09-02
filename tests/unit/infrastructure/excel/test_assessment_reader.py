from datetime import date, datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook

from lead_enrichment.infrastructure.excel.assessment_reader import (
    AssessmentWorkbookError,
    MissingAssessmentColumnsError,
    read_assessment_workbook,
)
from lead_enrichment.models import ChannelScope, ChannelType, ContactRole, LeadPriority

MAIN_HEADERS = [
    "Застройщик / бренд",
    "Регион",
    "Дата ассесмента",
    "Дата рейтинга источника",
    "ID ЕРЗ",
    "Ссылка на карточку ЕРЗ",
    "Сайт компании / проекта",
    "Активные ЖК по ЕРЗ",
    "Города активных ЖК по ЕРЗ",
    "Адрес по ЕРЗ",
    "Телефон отдела продаж по ЕРЗ",
    "VK по ЕРЗ",
    "Тип масштаба застройщика",
    "Итог Stage 1",
    "Статус совпадения с Indigo",
    "Совпавшая группа Indigo",
    "Совпавший алиас / проект",
    "Тип совпадения Indigo",
    "Статус текущей проработки",
    "Ответственный",
    "Дата последнего касания",
    "Баллы: коммерческий потенциал",
    "Баллы: маркетинговая потребность",
    "Баллы: digital-возможность",
    "Баллы: outreach-доступность",
    "Итоговый балл",
    "TIR",
    "Примененное cap-правило",
    "Зацепка для первого касания",
    "Следующее действие",
    "Недостающие данные",
    "Ссылки на источники",
    "Комментарии",
    "Источники ЛПР",
]

LPR_HEADERS = [
    "Строка",
    "Компания",
    "Проекты / ЖК",
    "Основной актуальный ЛПР",
    "Роль",
    "Альтернативный / стратегический ЛПР",
    "Источник / подтверждение",
    "Статус актуальности",
    "Рекомендация для outreach",
    "Комментарий",
    "Ссылки",
]


def _save_assessment(
    path: Path,
    *,
    main_rows: list[list[object]],
    lpr_rows: list[list[object]],
    main_headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "Ростовская область"
    main.append(main_headers or MAIN_HEADERS)
    for row in main_rows:
        main.append(row)
    lpr = workbook.create_sheet("ЛПР 2026 verified")
    lpr.append(LPR_HEADERS)
    for row in lpr_rows:
        lpr.append(row)
    workbook.save(path)
    workbook.close()


def test_reader_imports_assessment_context_channels_and_lpr(tmp_path: Path) -> None:
    path = tmp_path / "assessment.xlsx"
    _save_assessment(
        path,
        main_rows=[[
            "ГК Тест",
            "Ростовская область",
            "2026-07-12",
            "01.07.2026",
            12345,
            "https://erzrf.ru/company/12345",
            "example.ru",
            "ЖК Первый; ЖК Второй",
            "Ростов-на-Дону",
            "Тестовый адрес",
            "(863) 1234567, (863) 7654321",
            "https://vk.com/example",
            "regional",
            "REVIEW",
            "company_match",
            "Indigo group",
            "Тестовый алиас",
            "company",
            "check_required",
            None,
            None,
            30,
            20,
            10,
            15,
            75,
            "TIR 1",
            "none",
            "Персональная зацепка",
            "manual_check_before_outreach",
            "email ЛПР",
            "https://erzrf.ru/company/12345; https://example.ru/about",
            "Комментарий",
            "https://companies.rbc.ru/example",
        ]],
        lpr_rows=[[
            2,
            "ГК Тест",
            "ЖК Первый; ЖК Второй",
            "Иванов Иван Иванович",
            "директор по маркетингу",
            "Петров Пётр Петрович — коммерческий директор; Сидоров Сидор Сидорович — собственник",
            "Подтверждено публичным источником",
            "актуально: официальный сайт",
            "direct_top_lpr",
            "Проверить персональный email",
            "https://example.ru/team; https://media.example/person",
        ]],
    )

    result = read_assessment_workbook(
        path,
        collected_at=datetime(2026, 9, 2, tzinfo=timezone.utc),
    )

    assert result.summary.imported_rows == 1
    assert result.summary.rows_with_website == 1
    assert result.summary.rows_with_sales_phones == 1
    assert result.summary.primary_contacts == 1
    assert result.summary.alternative_contacts == 2
    assert result.summary.indigo_matches == 1
    assert result.summary.tier_counts == {"TIR 1": 1}

    company = result.companies[0]
    assert company.company_key == "assessment:erz:12345"
    assert company.assessment_date == date(2026, 7, 12)
    assert company.source_rating_date == date(2026, 7, 1)
    assert company.website == "https://example.ru"
    assert company.projects == ["ЖК Первый", "ЖК Второй"]
    assert company.lead_priority == LeadPriority.TIR_1
    assert company.scores.total == 75
    assert [channel.value for channel in company.company_channels] == [
        "+78631234567",
        "+78637654321",
        "https://vk.com/example",
    ]
    assert company.company_channels[0].scope == ChannelScope.COMPANY
    assert company.company_channels[-1].channel_type == ChannelType.SOCIAL
    assert [contact.normalized_role for contact in company.contacts] == [
        ContactRole.MARKETING,
        ContactRole.SALES,
        ContactRole.OWNER,
    ]
    assert all(contact.source_refs for contact in company.contacts)
    assert len(company.source_fields) == len(MAIN_HEADERS)


def test_reader_does_not_attach_lpr_when_company_reference_mismatches(tmp_path: Path) -> None:
    path = tmp_path / "mismatch.xlsx"
    _save_assessment(
        path,
        main_rows=[["ГК Тест"]],
        lpr_rows=[[2, "Другая компания", None, "Иванов Иван", "директор"]],
    )

    result = read_assessment_workbook(path)

    assert result.summary.imported_rows == 1
    assert result.companies[0].contacts == []
    assert [issue.code for issue in result.summary.issues] == ["LPR_COMPANY_MISMATCH"]


def test_reader_requires_main_and_lpr_sheets(tmp_path: Path) -> None:
    path = tmp_path / "missing-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "Ростовская область"
    workbook.active.append(["Застройщик / бренд"])
    workbook.save(path)
    workbook.close()

    with pytest.raises(AssessmentWorkbookError, match="ЛПР 2026 verified"):
        read_assessment_workbook(path)


def test_reader_reports_missing_required_columns(tmp_path: Path) -> None:
    path = tmp_path / "missing-columns.xlsx"
    _save_assessment(
        path,
        main_rows=[["Ростовская область"]],
        lpr_rows=[[2, "ГК Тест", None, "Иванов Иван", "директор"]],
        main_headers=["Регион"],
    )

    with pytest.raises(MissingAssessmentColumnsError) as error:
        read_assessment_workbook(path)

    assert error.value.sheet_name == "Ростовская область"
    assert error.value.missing_columns == ["Застройщик / бренд"]
