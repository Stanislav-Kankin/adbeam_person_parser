from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from lead_enrichment.infrastructure.excel.kontur_reader import (
    MissingRequiredColumnsError,
    read_kontur_workbook,
)
from lead_enrichment.models import ChannelType, ContactRole, EntityType

HEADERS = [
    "Наименование",
    "ИНН",
    "КПП",
    "ОГРН",
    "Дата регистрации",
    "ФИО руководителя",
    "Должность руководителя",
    "Номер телефона",
    "Электронная почта",
    "Карточка в Фокусе",
    "Ссылка на сайт",
    "Количество сотрудников",
    "Источник",
    "Название сегмента",
]


def _save_workbook(path: Path, rows: list[list[object]], headers: list[str] | None = None) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Контрагенты"
    worksheet.append(headers or HEADERS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def test_reader_imports_company_channels_manager_and_hyperlinks(tmp_path: Path) -> None:
    path = tmp_path / "kontur.xlsx"
    _save_workbook(
        path,
        [[
            'ООО "Тест"',
            "1234567894",
            "123456789",
            "1234567890123",
            datetime(2020, 1, 2),
            "Иванов Иван Иванович",
            "Генеральный директор",
            "+7 (999) 111-22-33\n8 999 111 22 33",
            "Info@Example.ru\nowner@example.org",
            "Открыть в Фокусе",
            "Открыть сайт",
            12,
            "Контур.Продажи",
            "Тестовый сегмент",
        ]],
    )
    workbook = load_workbook(path)
    worksheet = workbook["Контрагенты"]
    worksheet["J2"].hyperlink = "https://focus.example/company/1234567894"
    worksheet["K2"].hyperlink = "http://example.ru/"
    workbook.save(path)
    workbook.close()

    result = read_kontur_workbook(
        path,
        collected_at=datetime(2026, 9, 2, tzinfo=timezone.utc),
    )

    assert result.summary.imported_rows == 1
    company = result.companies[0]
    assert company.entity_type == EntityType.LEGAL_ENTITY
    assert company.focus_url == "https://focus.example/company/1234567894"
    assert company.website == "http://example.ru/"
    assert company.employee_count == 12
    assert [channel.channel_type for channel in company.company_channels] == [
        ChannelType.EMAIL,
        ChannelType.EMAIL,
        ChannelType.PHONE,
    ]
    assert company.initial_people[0].normalized_role == ContactRole.LEADER
    assert not hasattr(company.initial_people[0], "emails")


def test_reader_derives_owner_for_individual_entrepreneur(tmp_path: Path) -> None:
    path = tmp_path / "kontur-ip.xlsx"
    _save_workbook(
        path,
        [[
            "ИП Петров Пётр Петрович",
            "123456789047",
            None,
            "123456789012345",
            datetime(2021, 5, 6),
            None,
            None,
            None,
            "ip@example.ru",
            None,
            None,
            None,
            "Контур.Продажи",
            "Тестовый сегмент",
        ]],
    )

    result = read_kontur_workbook(path)

    company = result.companies[0]
    assert company.entity_type == EntityType.INDIVIDUAL_ENTREPRENEUR
    assert company.initial_people[0].full_name == "Петров Пётр Петрович"
    assert company.initial_people[0].normalized_role == ContactRole.OWNER
    assert company.company_channels[0].value == "ip@example.ru"


def test_reader_reports_invalid_inn_without_exposing_cell_value(tmp_path: Path) -> None:
    path = tmp_path / "invalid.xlsx"
    _save_workbook(
        path,
        [["ООО Тест", "1234567890"]],
        headers=["Наименование", "ИНН"],
    )

    result = read_kontur_workbook(path)

    assert result.summary.imported_rows == 0
    assert result.summary.skipped_rows == 1
    assert result.summary.issues[0].code == "INVALID_INN"
    assert "1234567890" not in result.summary.issues[0].message


def test_reader_requires_name_and_inn_columns(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"
    _save_workbook(path, [["ООО Тест"]], headers=["Наименование"])

    with pytest.raises(MissingRequiredColumnsError) as error:
        read_kontur_workbook(path)

    assert error.value.missing_columns == ["ИНН"]
