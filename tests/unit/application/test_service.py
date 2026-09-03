from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import lead_enrichment.application.service as service
from lead_enrichment.application import (
    PipelineRunRequest,
    inspect_kontur_input,
    run_kontur_pipeline,
)
from lead_enrichment.application.service import _validate_request
from lead_enrichment.models import SourceApplicability, SourceMetadata


class _FakeHttpClient:
    def __enter__(self):
        return self

    def __exit__(self, *_args) -> None:
        return None


class _SkippedSitePlugin:
    def __init__(self, _client) -> None:
        pass

    @property
    def metadata(self) -> SourceMetadata:
        return SourceMetadata(
            source_id="test_site",
            version="1.0.0",
            display_name="Test site",
        )

    def is_applicable(self, _context) -> SourceApplicability:
        return SourceApplicability(
            applicable=False,
            reason_code="TEST_SKIP",
            reason_message="Тестовый источник пропущен",
        )

    def execute(self, _context):
        raise AssertionError("execute must not be called")


def _save_kontur(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Контрагенты"
    sheet.append([
        "Наименование",
        "ИНН",
        "Ссылка на сайт",
        "ФИО руководителя",
        "Номер телефона",
    ])
    sheet.append(['ООО "Тест"', "1234567894", "example.ru", "Иванов Иван", "+7 495 111-22-33"])
    sheet.append(['ООО "Без ИНН"', None, "invalid.example", None, None])
    workbook.save(path)
    workbook.close()


def test_inspect_kontur_input_counts_only_valid_inn_rows(tmp_path: Path) -> None:
    path = tmp_path / "kontur.xlsx"
    _save_kontur(path)

    result = inspect_kontur_input(path)

    assert result.total_rows == 2
    assert result.imported_companies == 1
    assert result.skipped_rows == 1
    assert result.invalid_inn_rows == 1
    assert result.rows_with_website == 1
    assert result.rows_with_manager == 1
    assert result.rows_with_phones == 1


def test_pipeline_never_overwrites_kontur_input(tmp_path: Path) -> None:
    path = tmp_path / "kontur.xlsx"
    request = PipelineRunRequest(
        kontur_file=path,
        output_file=path,
        checkpoint_file=tmp_path / "state.sqlite3",
    )

    with pytest.raises(ValueError, match="перезаписывать"):
        _validate_request(request)


def test_run_kontur_pipeline_exports_inn_first_result(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_path = tmp_path / "kontur.xlsx"
    output_path = tmp_path / "result.xlsx"
    _save_kontur(input_path)
    monkeypatch.setattr(service, "HttpClient", _FakeHttpClient)
    monkeypatch.setattr(service, "SiteCrawlPlugin", _SkippedSitePlugin)

    result = run_kontur_pipeline(
        PipelineRunRequest(
            kontur_file=input_path,
            output_file=output_path,
            checkpoint_file=tmp_path / "state.sqlite3",
        )
    )

    assert result.run_id.startswith("kontur-")
    assert result.summary.total_companies == 1
    assert result.leads[0].company.inn == "1234567894"
    assert result.leads[0].lead.assessment is None
    assert output_path.is_file()
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        companies = workbook["Компании"]
        assert companies["B1"].value == "Наименование"
        assert companies["C2"].value == "1234567894"
        assert companies["T2"].value == "INN"
        assert workbook["Исходные данные"]["D2"].value == "1234567894"
    finally:
        workbook.close()
