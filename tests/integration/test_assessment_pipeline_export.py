from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from lead_enrichment.engine import EnrichmentOrchestrator, merge_assessment_with_kontur
from lead_enrichment.infrastructure import CheckpointRegistry
from lead_enrichment.infrastructure.excel import (
    export_batch_enrichment_workbook,
    read_assessment_workbook,
    read_kontur_workbook,
)
from lead_enrichment.models import (
    ChannelScope,
    ChannelType,
    ContactChannel,
    ContactRole,
    PersonContact,
    SourceApplicability,
    SourceMetadata,
    SourceOutcome,
    SourceResult,
)

NOW = datetime(2026, 9, 2, tzinfo=timezone.utc)


class _SyntheticPersonSource:
    @property
    def metadata(self) -> SourceMetadata:
        return SourceMetadata(
            source_id="synthetic_people",
            version="1.0.0",
            display_name="Synthetic people",
        )

    def is_applicable(self, _context) -> SourceApplicability:
        return SourceApplicability(
            applicable=True,
            reason_code="APPLICABLE",
            reason_message="Применим",
        )

    def execute(self, context) -> SourceResult:
        if context.company.brand_name != "Альфа":
            return SourceResult(
                source_id="synthetic_people",
                source_version="1.0.0",
                outcome=SourceOutcome.NOT_FOUND,
                reason_code="PERSON_NOT_FOUND",
                reason_message="Контакт не найден",
            )
        return SourceResult(
            source_id="synthetic_people",
            source_version="1.0.0",
            outcome=SourceOutcome.FOUND,
            reason_code="PERSON_FOUND",
            reason_message="Контакт найден",
            person_contacts=[
                PersonContact(
                    contact_id=uuid4(),
                    full_name="Иванов Иван",
                    job_title="Коммерческий директор",
                    normalized_role=ContactRole.SALES,
                    confidence_score=85,
                    channels=[
                        ContactChannel(
                            channel_type=ChannelType.EMAIL,
                            value="person@example.ru",
                            scope=ChannelScope.PERSONAL,
                        )
                    ],
                )
            ],
        )


def _assessment_file(path: Path) -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "Ростовская область"
    main.append([
        "Застройщик / бренд",
        "Регион",
        "ID ЕРЗ",
        "Сайт компании / проекта",
        "TIR",
        "Итоговый балл",
    ])
    main.append(["Альфа", "Ростовская область", "1", "alpha.ru", "TIR 1", 80])
    main.append(["Бета", "Ростовская область", "2", "beta.ru", "TIR 2", 60])
    lpr = workbook.create_sheet("ЛПР 2026 verified")
    lpr.append(["Строка", "Компания", "Основной актуальный ЛПР", "Роль"])
    lpr.append([2, "Альфа", None, None])
    lpr.append([3, "Бета", None, None])
    workbook.save(path)
    workbook.close()


def _kontur_file(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Контрагенты"
    sheet.append(["Наименование", "ИНН", "Ссылка на сайт"])
    sheet.append(['ООО "Альфа"', "1234567894", "alpha.ru"])
    sheet.append(['ООО "Бета"', "7707083893", "beta.ru"])
    workbook.save(path)
    workbook.close()


def test_full_assessment_merge_orchestrator_and_export(tmp_path: Path) -> None:
    assessment_path = tmp_path / "assessment.xlsx"
    kontur_path = tmp_path / "kontur.xlsx"
    output_path = tmp_path / "result.xlsx"
    _assessment_file(assessment_path)
    _kontur_file(kontur_path)

    assessment = read_assessment_workbook(assessment_path, collected_at=NOW)
    kontur = read_kontur_workbook(kontur_path, collected_at=NOW)
    merge = merge_assessment_with_kontur(assessment, kontur)
    batch = EnrichmentOrchestrator(
        [_SyntheticPersonSource()],
        checkpoint_registry=CheckpointRegistry(tmp_path / "state.sqlite3"),
    ).run(merge.leads, run_id="integration-run", collected_at=NOW)

    export_batch_enrichment_workbook(batch, output_path)

    assert batch.summary.resolved == 1
    assert batch.summary.manual_required == 1
    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook.sheetnames == [
            "Сводка",
            "Компании",
            "Контакты",
            "Источники",
            "Журнал",
            "Ручная очередь",
            "Исходные данные",
        ]
        assert workbook["Сводка"]["B6"].value == "=COUNTA('Компании'!A2:A3)"
        assert workbook["Компании"]["L2"].value == "RESOLVED"
        assert workbook["Компании"]["L3"].value == "MANUAL_REQUIRED"
        assert workbook["Контакты"]["D2"].value == "Иванов Иван"
        assert workbook["Контакты"]["H2"].value == "person@example.ru"
        assert workbook["Ручная очередь"].max_row == 5
        assert workbook["Исходные данные"].max_row == 3
    finally:
        workbook.close()
