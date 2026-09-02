from datetime import datetime, timezone
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook

from lead_enrichment.infrastructure.excel import (
    export_enrichment_workbook,
    read_kontur_workbook,
)
from lead_enrichment.infrastructure.http import HttpClient
from lead_enrichment.models import HttpClientSettings, PipelineContext, SiteCrawlSettings
from lead_enrichment.sources import SiteCrawlPlugin


def test_kontur_to_mocked_site_to_atomic_excel(tmp_path: Path) -> None:
    source_path = tmp_path / "kontur.xlsx"
    output_path = tmp_path / "enriched.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Контрагенты"
    sheet.append(["Наименование", "ИНН", "Ссылка на сайт", "Источник"])
    sheet.append(['ООО "Тест"', "1234567894", "https://example.com/", "Контур.Продажи"])
    workbook.save(source_path)
    workbook.close()

    imported = read_kontur_workbook(
        source_path,
        collected_at=datetime(2026, 9, 2, tzinfo=timezone.utc),
    )

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == "/robots.txt":
            return httpx.Response(404)
        return httpx.Response(
            200,
            headers={"Content-Type": "text/html"},
            text='<html><body><a href="mailto:hello@brand.ru">Email</a></body></html>',
        )

    with HttpClient(
        HttpClientSettings(
            max_attempts=1,
            retry_min_seconds=0,
            retry_max_seconds=0,
            per_host_delay_seconds=0,
        ),
        transport=httpx.MockTransport(handler),
        resolver=lambda _hostname: ["93.184.216.34"],
    ) as client:
        plugin = SiteCrawlPlugin(client, SiteCrawlSettings(max_pages=1))
        site_result = plugin.execute(
            PipelineContext(
                run_id="integration-test",
                company=imported.companies[0],
                collected_at=datetime(2026, 9, 2, tzinfo=timezone.utc),
            )
        )

    export_enrichment_workbook(
        imported,
        output_path,
        site_results={imported.companies[0].inn: site_result},
    )

    result = load_workbook(output_path, read_only=True, data_only=False)
    try:
        assert result["Контакты"]["F2"].value == "hello@brand.ru"
        assert result["Журнал"]["C3"].value == "FOUND"
        assert result["Журнал"]["D3"].value == "SITE_DIRECT_CONTACTS_FOUND"
    finally:
        result.close()
